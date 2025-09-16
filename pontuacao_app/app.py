# sistema_pontuacao_flask.py

from flask import Flask, render_template, request, redirect, flash, url_for
from datetime import datetime
from decimal import Decimal, ROUND_HALF_UP
import psycopg2
import os
import re
import logging
import cloudinary
import cloudinary.uploader
import tempfile
import zipfile
import pandas as pd
from flask import send_file
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
DELETE_PASSWORD = 'confie123'

#logs
logger = logging.getLogger(__name__)
if not logger.handlers:
    logging.basicConfig(level=logging.INFO)


# --- ResponsÃ¡veis por critÃ©rio (Aâ€“E) por setor/tabela ---
RESPONSABILIDADES = {
    'loja': {
        'GERENTE_ADM': ['A', 'E'],
        'RH':          ['B', 'D'],
        'FINANCEIRO':  ['C'],
        # Faturamento nÃ£o atua na Loja
    },
    'expedicao': {
        'GERENTE_ADM': ['A'],
        'FATURAMENTO': ['B'],
        'FINANCEIRO':  ['C', 'D'],
        'RH':          ['E'],
    },
    'logistica': {
        'FINANCEIRO':  ['A', 'C'],
        'GERENTE_ADM': ['B'],
        'RH':          ['D'],
        'FATURAMENTO': ['E'],
    },
    'comercial': {
        'SUPERVISOR':        ['A'],
        'GERENTE_COMERCIAL': ['B'],
        'FATURAMENTO':       ['C'],
        'FINANCEIRO':        ['D'],
        'RH':                ['E'],
    }
}

def _filtro_responsavel_sql(tabela: str, responsavel: str):
    """
    Monta o trecho de WHERE para filtrar registros onde *esse responsÃ¡vel atuou*.
    Regra: atuou se QUALQUER coluna atribuÃ­da a ele for diferente de 0 (<> 0).
    Retorna (sql_fragment, params). Ex.: ("(A <> 0 OR C <> 0)", [])
    """
    if not responsavel:
        return "", []
    tab = tabela.lower().strip()
    resp = responsavel.strip().upper()
    cols = RESPONSABILIDADES.get(tab, {}).get(resp)
    if not cols:
        return "", []
    conds = [f"{c} <> 0" for c in cols]
    return "(" + " OR ".join(conds) + ")", []



app = Flask(__name__)
app.secret_key = 'confie'

# ConexÃ£o com o banco PostgreSQL no Render
def get_db_connection():
    return psycopg2.connect(os.environ['DATABASE_URL'])

# Inicializa o banco e cria as tabelas se nÃ£o existirem
def init_db():
    conn = get_db_connection()
    c = conn.cursor()

    c.execute('''
        CREATE TABLE IF NOT EXISTS pontuacoes (
            id SERIAL PRIMARY KEY,
            data TEXT,
            setor TEXT,
            obrigacao TEXT,
            pontuacao TEXT,
            observacao TEXT
        )
    ''')
    c.execute('''
        CREATE TABLE IF NOT EXISTS loja (
            id SERIAL PRIMARY KEY,
            data TEXT,
            A INTEGER,
            B INTEGER,
            C INTEGER,
            D INTEGER,
            E INTEGER,
            extras TEXT,
            observacao TEXT,
            total INTEGER
        )
    ''')
    c.execute('''
        CREATE TABLE IF NOT EXISTS expedicao (
            id SERIAL PRIMARY KEY,
            data TEXT,
            A INTEGER,
            B INTEGER,
            C INTEGER,
            D INTEGER,
            E INTEGER,
            extras TEXT,
            observacao TEXT,
            total INTEGER
        )
    ''')
    c.execute('''
        CREATE TABLE IF NOT EXISTS logistica (
            id SERIAL PRIMARY KEY,
            data TEXT,
            motorista TEXT,
            A INTEGER,
            B INTEGER,
            C INTEGER,
            D INTEGER,
            E INTEGER,
            extras TEXT,
            observacao TEXT,
            total INTEGER
        )
    ''')
    c.execute('''
        CREATE TABLE IF NOT EXISTS comercial (
            id SERIAL PRIMARY KEY,
            data TEXT,
            vendedor TEXT,
            A INTEGER,
            B INTEGER,
            C INTEGER,
            D INTEGER,
            E INTEGER,
            extras TEXT,
            observacao TEXT,
            total INTEGER
        )
    ''')

    conn.commit()
    conn.close()

cloudinary.config()

def fazer_backup_e_enviar():
    conn = None
    try:
        conn = get_db_connection()
        c = conn.cursor()

        tabelas = ['loja', 'expedicao', 'logistica', 'comercial', 'pontuacoes']
        arquivos_csv = []

        with tempfile.TemporaryDirectory() as tmpdirname:
            logger.info(f"[Backup] Pasta temporÃ¡ria criada: {tmpdirname}")

            for tabela in tabelas:
                c.execute(f"SELECT * FROM {tabela}")
                rows = c.fetchall()
                colnames = [desc[0] for desc in c.description]

                if not rows:
                    logger.info(f"[Backup] tabela '{tabela}' vazia â€” pulando")
                    continue

                df = pd.DataFrame(rows, columns=colnames)
                caminho_csv = os.path.join(tmpdirname, f"{tabela}.csv")
                df.to_csv(caminho_csv, index=False)
                arquivos_csv.append(caminho_csv)
                logger.info(f"[Backup] gerado CSV: {caminho_csv} ({len(rows)} linhas)")

            if not arquivos_csv:
                logger.info("[Backup] Nenhuma tabela com dados â€” nada para enviar")
                return None

            caminho_zip = os.path.join(tmpdirname, f"backup_{datetime.now().strftime('%Y-%m-%d_%H-%M-%S')}.zip")
            with zipfile.ZipFile(caminho_zip, 'w') as zipf:
                for file in arquivos_csv:
                    zipf.write(file, os.path.basename(file))
            logger.info(f"[Backup] ZIP criado em: {caminho_zip} ({os.path.getsize(caminho_zip)} bytes)")

            public_id = f"pontuacao_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
            max_retries = 3
            attempt = 0
            last_exc = None

            while attempt < max_retries:
                attempt += 1
                try:
                    logger.info(f"[Backup] Iniciando upload (attempt {attempt}) public_id={public_id}")
                    resultado = cloudinary.uploader.upload(
                        caminho_zip,
                        resource_type='raw',
                        folder='backups_pontuacao',
                        use_filename=True,
                        unique_filename=True,   # evita conflito de nomes
                        overwrite=False,
                        public_id=public_id
                    )
                    logger.info(f"[Backup] Upload OK: {resultado.get('secure_url')} (public_id={resultado.get('public_id')})")
                    return resultado.get('secure_url')
                except Exception as e:
                    last_exc = e
                    logger.exception(f"[Backup] Erro no upload attempt={attempt}: {e}")
            logger.error(f"[Backup] Falhou apÃ³s {max_retries} tentativas. Erro final: {last_exc}")
            return None

    except Exception as e:
        logger.exception("Erro ao fazer backup automÃ¡tico (outer): %s", e)
        return None
    finally:
        try:
            if conn:
                conn.close()
        except Exception:
            pass

# Conversor de data (dd/mm/aaaa ou yyyy-mm-dd â†’ yyyy-mm-dd)
def norm_date_to_iso(s):
    for fmt in ("%Y-%m-%d", "%d/%m/%Y"):
        try:
            return datetime.strptime(s.strip(), fmt).strftime("%Y-%m-%d")
        except Exception:
            pass
    return None


@app.template_filter('datetimeformat')
def datetimeformat(value, format='%d/%m/%Y'):
    try:
        return datetime.strptime(value, "%Y-%m-%d").strftime(format)
    except:
        return value

@app.route('/', endpoint='home_pontuacao')
def home():
    conn = get_db_connection()
    c = conn.cursor()

    setores = []

    # Loja (sem mÃ©dia)
    c.execute("SELECT total FROM loja")
    loja_pontos = [row[0] for row in c.fetchall()]
    setores.append({
        'nome': 'Loja',
        'total_registros': len(loja_pontos),
        'soma': sum(loja_pontos) if loja_pontos else 0,
        'media': None,
        'valor_grafico': sum(loja_pontos) if loja_pontos else 0
    })

    # ExpediÃ§Ã£o (sem mÃ©dia)
    c.execute("SELECT total FROM expedicao")
    expedicao_pontos = [row[0] for row in c.fetchall()]
    setores.append({
        'nome': 'ExpediÃ§Ã£o',
        'total_registros': len(expedicao_pontos),
        'soma': sum(expedicao_pontos) if expedicao_pontos else 0,
        'media': None,
        'valor_grafico': sum(expedicao_pontos) if expedicao_pontos else 0
    })

    # LogÃ­stica (usa mÃ©dia)
    c.execute("SELECT total FROM logistica")
    logistica_pontos = [row[0] for row in c.fetchall()]
    soma_log = sum(logistica_pontos) if logistica_pontos else 0
    media_log = round(soma_log / 6, 2) if soma_log else 0
    setores.append({
        'nome': 'LogÃ­stica',
        'total_registros': len(logistica_pontos),
        'soma': soma_log,
        'media': media_log,
        'valor_grafico': media_log
    })

    # Comercial (usa mÃ©dia)
    c.execute("SELECT total FROM comercial")
    comercial_pontos = [row[0] for row in c.fetchall()]
    soma_com = sum(comercial_pontos) if comercial_pontos else 0
    media_com = round(soma_com / 8, 2) if soma_com else 0
    setores.append({
        'nome': 'Comercial',
        'total_registros': len(comercial_pontos),
        'soma': soma_com,
        'media': media_com,
        'valor_grafico': media_com
    })

    c.close()
    conn.close()

    return render_template(
        'home.html',
        total_loja=setores[0]['total_registros'],
        soma_loja=setores[0]['soma'],
        total_expedicao=setores[1]['total_registros'],
        soma_expedicao=setores[1]['soma'],
        total_logistica=setores[2]['total_registros'],
        soma_logistica=setores[2]['soma'],
        media_logistica=setores[2]['media'],
        total_comercial=setores[3]['total_registros'],
        soma_comercial=setores[3]['soma'],
        media_comercial=setores[3]['media']
    )


@app.route('/enviar', methods=['POST'])
def enviar():
    setor = request.form['setor']
    obrigacao = request.form['obrigacao']
    pontuacao = request.form['pontuacao']
    observacao = request.form.get('observacao', '')
    data = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    conn = get_db_connection()
    c = conn.cursor()

    c.execute('INSERT INTO pontuacoes (data, setor, obrigacao, pontuacao, observacao) VALUES (%s, %s, %s, %s, %s)',
              (data, setor, obrigacao, pontuacao, observacao))
    conn.commit()
    conn.close()

    flash("âœ… PontuaÃ§Ã£o registrada com sucesso!", "success")
    return redirect(url_for('home_pontuacao'))


@app.route('/historico')
def historico():
    conn = get_db_connection()
    c = conn.cursor()
    c.execute('SELECT data, setor, obrigacao, pontuacao, observacao FROM pontuacoes ORDER BY data DESC')
    registros = c.fetchall()
    conn.close()
    return render_template('historico.html', registros=registros)

# =======================================================================
# LOJA
# =======================================================================
# FunÃ§Ã£o para converter valores de forma segura
def safe_int(value):
    try:
        return int(value)
    except (ValueError, TypeError):
        return 0

# Dentro da rota de envio (exemplo para Loja)
@app.route('/loja', methods=['GET', 'POST'])
def loja():
    if request.method == 'POST':
        data_unica = request.form.get('data', '').strip()
        criterios  = request.form.getlist('criterios')
        observacao = request.form.get('observacao', '')
        extras     = request.form.getlist('extras')

        pesos = {'A': 1, 'B': 1, 'C': 1, 'D': -1, 'E': -2}

        A = int('A' in criterios)
        B = int('B' in criterios)
        C = int('C' in criterios)
        D = int('D' in criterios)
        E = int('E' in criterios)

        datas_raw = request.form.get('datas', '').strip()
        lista_datas, invalidas = [], []

        if datas_raw:
            tokens = re.split(r'[,\n;\s]+', datas_raw)
            for t in tokens:
                if not t:
                    continue
                iso = norm_date_to_iso(t)
                if iso:
                    lista_datas.append(iso)
                else:
                    invalidas.append(t)

        if not lista_datas:
            iso = norm_date_to_iso(data_unica or '')
            if not iso:
                flash('âŒ Informe a data ou selecione mÃºltiplas datas no formato dd/mm/aaaa.', 'danger')
                return redirect(url_for('loja'))
            lista_datas = [iso]

        lista_datas = sorted(set(lista_datas))

        inseridos = 0
        pulados   = 0

        conn = get_db_connection()
        c = conn.cursor()
        try:
            for dia in lista_datas:
                c.execute("SELECT A, B, C, D, E FROM loja WHERE data = %s", (dia,))
                registros = c.fetchall()

                conflito = False
                for registro in registros:
                    crits_existentes = {'A': registro[0], 'B': registro[1], 'C': registro[2], 'D': registro[3], 'E': registro[4]}
                    for c_sel in criterios:
                        if crits_existentes.get(c_sel, 0) == 1:
                            conflito = True
                            break
                    if conflito:
                        break

                if conflito:
                    pulados += 1
                    continue

                total = sum([pesos[c] for c in criterios])
                if 'meta' in extras:
                    total += 2
                if 'equipe90' in extras:
                    total += 1

                c.execute("""
                    INSERT INTO loja (data, A, B, C, D, E, extras, observacao, total)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
                """, (dia, A, B, C, D, E, ','.join(extras), observacao, total))
                inseridos += 1

            conn.commit()
        finally:
            conn.close()

        msgs = []
        if inseridos: msgs.append(f"âœ… {inseridos} registro(s) inserido(s).")
        if pulados:   msgs.append(f"âš ï¸ {pulados} dia(s) pulado(s) por jÃ¡ conterem os mesmos critÃ©rios.")
        if invalidas: msgs.append(f"âŒ Datas invÃ¡lidas ignoradas: {', '.join(invalidas)}")

        flash(' '.join(msgs) if msgs else "Nada a fazer.", "success" if inseridos else "warning")
        fazer_backup_e_enviar()
        return redirect(url_for('loja'))

    # GET âžœ renderiza (sem redirect!)
    return render_template('loja.html')



# =======================================================================
# EXPEDIÃ‡ÃƒO
# =======================================================================
@app.route('/expedicao', methods=['GET', 'POST'])
def expedicao():
    if request.method == 'POST':
        data_unica = request.form.get('data', '').strip()
        criterios = request.form.getlist('criterios')

        A = 1  if 'A' in criterios else 0
        B = 1  if 'B' in criterios else 0
        C = 1  if 'C' in criterios else 0
        D = -2 if 'D' in criterios else 0
        E = -1 if 'E' in criterios else 0

        observacao = request.form.get('observacao', '')
        extras     = request.form.getlist('extras')

        extras_pontos = (2 if 'meta' in extras else 0) + (1 if 'equipe90' in extras else 0)
        total_base = A + B + C + D + E + extras_pontos

        datas_raw = request.form.get('datas', '').strip()
        lista_datas, invalidas = [], []

        if datas_raw:
            tokens = re.split(r'[,\n;\s]+', datas_raw)
            for t in tokens:
                if not t:
                    continue
                iso = norm_date_to_iso(t)
                if iso:
                    lista_datas.append(iso)
                else:
                    invalidas.append(t)

        if not lista_datas:
            iso = norm_date_to_iso(data_unica or '')
            if not iso:
                flash('âŒ Informe a data ou selecione mÃºltiplas datas no formato dd/mm/aaaa.', 'danger')
                return redirect(url_for('expedicao'))
            lista_datas = [iso]

        lista_datas = sorted(set(lista_datas))

        inseridos, pulados = 0, []

        conn = get_db_connection()
        c = conn.cursor()
        try:
            for dia in lista_datas:
                c.execute("SELECT A, B, C, D, E FROM expedicao WHERE data = %s", (dia,))
                registros_dia = c.fetchall()

                conflito = (
                    ('A' in criterios and any(r[0] == 1   for r in registros_dia)) or
                    ('B' in criterios and any(r[1] == 1   for r in registros_dia)) or
                    ('C' in criterios and any(r[2] == 1   for r in registros_dia)) or
                    ('D' in criterios and any(r[3] == -2  for r in registros_dia)) or
                    ('E' in criterios and any(r[4] == -1  for r in registros_dia))
                )
                if conflito:
                    pulados.append(dia)
                    continue

                c.execute("""
                    INSERT INTO expedicao (data, A, B, C, D, E, extras, observacao, total)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
                """, (dia, A, B, C, D, E, ','.join(extras), observacao, total_base))
                inseridos += 1

            conn.commit()
        finally:
            conn.close()

        msgs = []
        if inseridos: msgs.append(f"âœ… {inseridos} registro(s) inserido(s).")
        if pulados:   msgs.append(f"âš ï¸ Dias pulados por jÃ¡ conterem os mesmos critÃ©rios: {', '.join(pulados)}.")
        if invalidas: msgs.append(f"âŒ Datas invÃ¡lidas ignoradas: {', '.join(invalidas)}")

        flash(' '.join(msgs) if msgs else "Nada a fazer.", "success" if inseridos else "warning")
        fazer_backup_e_enviar()
        return redirect(url_for('expedicao'))

    # GET âžœ renderiza (sem redirect!)
    return render_template('expedicao.html')


@app.route('/historico_expedicao')
def historico_expedicao():
    responsavel = request.args.get('responsavel', '').strip()
    inicio = request.args.get('inicio', '').strip()
    fim    = request.args.get('fim', '').strip()

    conn = get_db_connection()
    c = conn.cursor()

    where, params = [], []

    if inicio:
        iso = norm_date_to_iso(inicio)
        if iso: where.append("data >= %s"); params.append(iso)
    if fim:
        iso = norm_date_to_iso(fim)
        if iso: where.append("data <= %s"); params.append(iso)

    sql_resp, p_resp = _filtro_responsavel_sql('expedicao', responsavel)
    if sql_resp:
        where.append(sql_resp)
        params += p_resp

    where_sql = (" WHERE " + " AND ".join(where)) if where else ""
    c.execute(f"""
        SELECT id, data, A, B, C, D, E, extras, total, observacao
        FROM expedicao{where_sql} ORDER BY data DESC
    """, params)
    registros = c.fetchall()
    conn.close()

    total_geral = sum([r[8] for r in registros]) if registros else 0

    return render_template('historico_expedicao.html',
                           registros=registros,
                           total_geral=total_geral,
                           responsavel=responsavel,
                           inicio=inicio,
                           fim=fim)

	

# Nova rota para a LogÃ­stica com menu de motoristas e formulÃ¡rio de pontuaÃ§Ã£o
@app.route('/logistica', methods=['GET', 'POST'])
def logistica():
    conn = get_db_connection()
    c = conn.cursor()

    c.execute('''
        CREATE TABLE IF NOT EXISTS logistica (
            id SERIAL PRIMARY KEY,
            data TEXT,
            motorista TEXT,
            A INTEGER,
            B INTEGER,
            C INTEGER,
            D INTEGER,
            E INTEGER,
            extras TEXT,
            observacao TEXT,
            total INTEGER
        )
    ''')

    motoristas = ['Denilson', 'Fabio', 'Rogerio', 'Robson', 'Simone', 'Vinicius', 'Equipe']

    if request.method == 'POST':
        data_unica = request.form.get('data', '').strip()
        motorista  = request.form['motorista']
        A = safe_int(request.form.get('A'))
        # B pode ter pontos customizados via campo B_valor
        if request.form.get("B") is not None:
            B = safe_int(request.form.get("B_valor"))
        else:
            B = 0
        C = safe_int(request.form.get('C'))
        D = safe_int(request.form.get('D'))
        E = safe_int(request.form.get('E'))
        extras = request.form.getlist('extras')
        observacao = request.form.get('observacao', '')

        # total base
        total = A + B + C + D + E

        # extra: economia (+2)
        if 'economia' in extras:
            total += 2

        # validaÃ§Ã£o do extra 'equipe90'
        if 'equipe90' in extras and motorista != 'Equipe':
            flash("âŒ O ponto extra 'Equipe chegou a 90%' sÃ³ pode ser usado com o motorista 'Equipe'.", "danger")
            conn.close()
            return redirect(url_for('logistica'))
        if 'equipe90' in extras:
            total += 1

        # === NOVO: mÃºltiplas datas (dd/mm/aaaa aceito) ===
        datas_raw = request.form.get('datas', '').strip()
        lista_datas, invalidas = [], []

        if datas_raw:
            import re
            tokens = re.split(r'[,\n;\s]+', datas_raw)  # vÃ­rgula, espaÃ§o ou quebra de linha
            for t in tokens:
                if not t:
                    continue
                iso = norm_date_to_iso(t)  # dd/mm/aaaa ou yyyy-mm-dd -> yyyy-mm-dd
                if iso:
                    lista_datas.append(iso)
                else:
                    invalidas.append(t)

        if not lista_datas:
            iso = norm_date_to_iso(data_unica or '')
            if not iso:
                flash('âŒ Informe a data ou selecione mÃºltiplas datas no formato dd/mm/aaaa.', 'danger')
                conn.close()
                return redirect(url_for('logistica'))
            lista_datas = [iso]

        # Evita datas duplicadas
        lista_datas = sorted(set(lista_datas))

        inseridos, pulados = 0, []

        try:
            for dia in lista_datas:
                # Travas por motorista + data (mesmo comportamento de antes)
                c.execute("SELECT A, B, C, D, E FROM logistica WHERE data = %s AND motorista = %s", (dia, motorista))
                registros = c.fetchall()

                conflito = (
                    (A == 1   and any(r[0] == 1   for r in registros)) or
                    (B == 1   and any(r[1] == 1   for r in registros)) or
                    (C == 1   and any(r[2] == 1   for r in registros)) or
                    (D == -2  and any(r[3] == -2  for r in registros)) or
                    (E == -1  and any(r[4] == -1  for r in registros))
                )
                if conflito:
                    pulados.append(dia)
                    continue

                # INSERT
                c.execute('''
                    INSERT INTO logistica (data, motorista, A, B, C, D, E, extras, observacao, total)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                ''', (dia, motorista, A, B, C, D, E, ','.join(extras), observacao, total))
                inseridos += 1

            conn.commit()
        finally:
            conn.close()

        # Feedback consolidado
        msgs = []
        if inseridos:
            msgs.append(f"âœ… {inseridos} registro(s) inserido(s).")
        if pulados:
            msgs.append(f"âš ï¸ Dias pulados por jÃ¡ conterem os mesmos critÃ©rios: {', '.join(pulados)}.")
        if invalidas:
            msgs.append(f"âŒ Datas invÃ¡lidas ignoradas: {', '.join(invalidas)}")

        flash(' '.join(msgs) if msgs else "Nada a fazer.", "success" if inseridos else "warning")
        fazer_backup_e_enviar()
        return redirect(url_for('logistica'))

    conn.close()
    return render_template('logistica.html', motoristas=motoristas)



@app.route('/historico_logistica')
def historico_logistica():
    motorista   = request.args.get('motorista', '').strip()
    responsavel = request.args.get('responsavel', '').strip()
    inicio      = request.args.get('inicio', '').strip()
    fim         = request.args.get('fim', '').strip()

    conn = get_db_connection()
    c = conn.cursor()

    motoristas = ['Denilson', 'Fabio', 'Rogerio', 'Robson', 'Simone', 'Vinicius', 'Equipe']

    where, params = [], []

    if motorista:
        where.append("motorista = %s"); params.append(motorista)

    if inicio:
        iso = norm_date_to_iso(inicio)
        if iso: where.append("data >= %s"); params.append(iso)
    if fim:
        iso = norm_date_to_iso(fim)
        if iso: where.append("data <= %s"); params.append(iso)

    sql_resp, p_resp = _filtro_responsavel_sql('logistica', responsavel)
    if sql_resp:
        where.append(sql_resp)
        params += p_resp

    where_sql = (" WHERE " + " AND ".join(where)) if where else ""
    c.execute(f"""
        SELECT id, data, motorista, A, B, C, D, E, extras, observacao, total
        FROM logistica{where_sql} ORDER BY data DESC
    """, params)
    registros = c.fetchall()
    conn.close()

    total_geral = sum([int(r[10]) for r in registros]) if registros else 0

    # mÃ©dia: mesma lÃ³gica que vocÃª jÃ¡ tinha, sÃ³ reaproveitei
    if motorista:
        from decimal import Decimal, ROUND_HALF_UP
        media = Decimal(total_geral).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
    else:
        from decimal import Decimal, ROUND_HALF_UP
        motoristas_reais = [m for m in motoristas if m != 'Equipe']
        media = (Decimal(total_geral / len(motoristas_reais)).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
                 if motoristas_reais else Decimal('0.00'))

    return render_template('historico_logistica.html',
                           registros=registros,
                           motorista=motorista,
                           motoristas=motoristas,
                           total_geral=total_geral,
                           media=media,
                           responsavel=responsavel,
                           inicio=inicio,
                           fim=fim)


@app.route('/historico_loja')
def historico_loja():
    responsavel = request.args.get('responsavel', '').strip()
    inicio = request.args.get('inicio', '').strip()
    fim    = request.args.get('fim', '').strip()

    conn = get_db_connection()
    c = conn.cursor()

    where = []
    params = []

    if inicio:
        iso = norm_date_to_iso(inicio)
        if iso: where.append("data >= %s"); params.append(iso)
    if fim:
        iso = norm_date_to_iso(fim)
        if iso: where.append("data <= %s"); params.append(iso)

    sql_resp, p_resp = _filtro_responsavel_sql('loja', responsavel)
    if sql_resp:
        where.append(sql_resp)
        params += p_resp

    where_sql = (" WHERE " + " AND ".join(where)) if where else ""
    c.execute(f"""
        SELECT id, data, A, B, C, D, E, extras, total, observacao
        FROM loja{where_sql} ORDER BY data DESC
    """, params)
    registros = c.fetchall()
    conn.close()

    total_geral = sum([r[8] for r in registros]) if registros else 0
    media = round(total_geral / len(registros), 1) if registros else 0

    return render_template('historico_loja.html',
                           registros=registros,
                           total_geral=total_geral,
                           media=media,
                           responsavel=responsavel,
                           inicio=inicio,
                           fim=fim)


# =======================================================================
# COMERCIAL
# =======================================================================
@app.route('/comercial', methods=['GET', 'POST'])
def comercial():
    conn = get_db_connection()
    c = conn.cursor()

    c.execute('''
        CREATE TABLE IF NOT EXISTS comercial (
            id SERIAL PRIMARY KEY,
            data TEXT,
            vendedor TEXT,
            A INTEGER,
            B INTEGER,
            C INTEGER,
            D INTEGER,
            E INTEGER,
            extras TEXT,
            observacao TEXT,
            total INTEGER
        )
    ''')

    vendedores = ['EVERTON', 'MARCELO', 'PEDRO', 'SILVANA', 'TIAGO', 'RODOLFO', 'MARCOS', 'THYAGO', 'EQUIPE']

    if request.method == 'POST':
        data = request.form['data']
        vendedor = request.form['vendedor']
        A = safe_int(request.form.get('A'))
        # B pode ter pontos customizados via campo B_valor
        if request.form.get("B") is not None:
            B = safe_int(request.form.get("B_valor"))
        else:
            B = 0
        C = safe_int(request.form.get('C'))
        D = safe_int(request.form.get('D'))
        E = safe_int(request.form.get('E'))
        extras = request.form.getlist('extras')
        observacao = request.form.get('observacao', '')

        # ðŸ”’ ValidaÃ§Ã£o do ponto extra equipe90
        if 'equipe90' in extras and vendedor.upper() != 'EQUIPE':
            flash("âŒ O ponto extra 'Equipe chegou a 90%' sÃ³ pode ser usado com o vendedor 'EQUIPE'.", "danger")
            conn.close()
            return redirect(url_for('comercial'))

        # ðŸ”’ Travas por vendedor e data (agora considerando qualquer valor diferente de zero)
        c.execute("SELECT A, B, C, D, E FROM comercial WHERE data = %s AND vendedor = %s", (data, vendedor))
        registros = c.fetchall()

        if A != 0 and any(r[0] != 0 for r in registros):
            flash("âš ï¸ O critÃ©rio A jÃ¡ foi registrado para esse vendedor nesse dia.", "danger")
            conn.close()
            return redirect(url_for('comercial'))
        if B != 0 and any(r[1] != 0 for r in registros):
            flash("âš ï¸ O critÃ©rio B jÃ¡ foi registrado para esse vendedor nesse dia.", "danger")
            conn.close()
            return redirect(url_for('comercial'))
        if C != 0 and any(r[2] != 0 for r in registros):
            flash("âš ï¸ O critÃ©rio C jÃ¡ foi registrado para esse vendedor nesse dia.", "danger")
            conn.close()
            return redirect(url_for('comercial'))
        if D != 0 and any(r[3] != 0 for r in registros):
            flash("âš ï¸ O critÃ©rio D jÃ¡ foi registrado para esse vendedor nesse dia.", "danger")
            conn.close()
            return redirect(url_for('comercial'))
        if E != 0 and any(r[4] != 0 for r in registros):
            flash("âš ï¸ O critÃ©rio E jÃ¡ foi registrado para esse vendedor nesse dia.", "danger")
            conn.close()
            return redirect(url_for('comercial'))

        total = A + B + C + D + E
        if 'meta' in extras:
            total += 2
        if 'equipe90' in extras:
            total += 1

        c.execute('''
            INSERT INTO comercial (data, vendedor, A, B, C, D, E, extras, observacao, total)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        ''', (data, vendedor, A, B, C, D, E, ','.join(extras), observacao, total))

        conn.commit()
        conn.close()

        flash("âœ… PontuaÃ§Ã£o registrada com sucesso!", "success")
        fazer_backup_e_enviar()
        return redirect(url_for('comercial'))

    conn.close()
    return render_template('comercial.html', vendedores=vendedores)


@app.route('/historico_comercial')

def historico_comercial():

Â  Â  vendedorÂ  Â  = request.args.get('vendedor', '').strip()

Â  Â  responsavel = request.args.get('responsavel', '').strip()

Â  Â  inicioÂ  Â  Â  = request.args.get('inicio', '').strip()

Â  Â  fimÂ  Â  Â  Â  Â = request.args.get('fim', '').strip()



Â  Â  conn = get_db_connection()

Â  Â  c = conn.cursor()



Â  Â  lista_vendedores = ['EVERTON', 'MARCELO', 'PEDRO', 'SILVANA', 'TIAGO', 'RODOLFO', 'MARCOS', 'THYAGO', 'EQUIPE']



Â  Â  where, params = [], []



Â  Â  if vendedor:

Â  Â  Â  Â  where.append("vendedor = %s"); params.append(vendedor)



Â  Â  if inicio:

Â  Â  Â  Â  iso = norm_date_to_iso(inicio)

Â  Â  Â  Â  if iso: where.append("data >= %s"); params.append(iso)

Â  Â  if fim:

Â  Â  Â  Â  iso = norm_date_to_iso(fim)

Â  Â  Â  Â  if iso: where.append("data <= %s"); params.append(iso)



Â  Â  sql_resp, p_resp = _filtro_responsavel_sql('comercial', responsavel)

Â  Â  if sql_resp:

Â  Â  Â  Â  where.append(sql_resp)

Â  Â  Â  Â  params += p_resp



Â  Â  where_sql = (" WHERE " + " AND ".join(where)) if where else ""

Â  Â  c.execute(f"""

Â  Â  Â  Â  SELECT * FROM comercial{where_sql} ORDER BY data DESC

Â  Â  """, params)

Â  Â  registros = c.fetchall()



Â  Â  total_geral = sum([r[10] for r in registros]) if registros else 0

Â  Â  from decimal import Decimal, ROUND_HALF_UP

Â  Â  media = Decimal(total_geral / 8).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP) if registros else Decimal('0.00')



Â  Â  conn.close()



Â  Â  return render_template('historico_comercial.html',

Â  Â  Â  Â  Â  Â  Â  Â  Â  Â  Â  Â  Â  Â registros=registros,

Â  Â  Â  Â  Â  Â  Â  Â  Â  Â  Â  Â  Â  Â total_geral=total_geral,

Â  Â  Â  Â  Â  Â  Â  Â  Â  Â  Â  Â  Â  Â media=media,

Â  Â  Â  Â  Â  Â  Â  Â  Â  Â  Â  Â  Â  Â vendedores=lista_vendedores,

Â  Â  Â  Â  Â  Â  Â  Â  Â  Â  Â  Â  Â  Â vendedor=vendedor,

Â  Â  Â  Â  Â  Â  Â  Â  Â  Â  Â  Â  Â  Â responsavel=responsavel,

Â  Â  Â  Â  Â  Â  Â  Â  Â  Â  Â  Â  Â  Â inicio=inicio,

Â  Â  Â  Â  Â  Â  Â  Â  Â  Â  Â  Â  Â  Â fim=fim)



@app.route('/zerar_tudo', methods=['POST'])
def zerar_tudo():
    senha = request.form.get('senha')
    if senha == "confie123":  # ajuste para sua senha desejada
        conn = get_db_connection()
        c = conn.cursor()
        for tabela in ['loja', 'expedicao', 'logistica', 'comercial']:
            c.execute(f"DELETE FROM {tabela}")
        conn.commit()
        conn.close()
        flash("âœ… Todas as pontuaÃ§Ãµes foram zeradas com sucesso!", "success")
    else:
        flash("âŒ Senha incorreta. AÃ§Ã£o cancelada.", "danger")

    return redirect(url_for('home_pontuacao'))


@app.route('/criar_banco')
def criar_banco():
    try:
        init_db()
        return "âœ… Banco de dados criado com sucesso!"
    except Exception as e:
        return f"âŒ Erro ao criar banco: {str(e)}"

@app.route('/restaurar_backup', methods=['GET', 'POST'])
def restaurar_backup():
    if request.method == 'POST':
        arquivo = request.files['backup']
        if arquivo and arquivo.filename.endswith('.zip'):
            try:
                with tempfile.TemporaryDirectory() as tmpdirname:
                    caminho_zip = os.path.join(tmpdirname, arquivo.filename)
                    arquivo.save(caminho_zip)

                    # Extrair ZIP
                    with zipfile.ZipFile(caminho_zip, 'r') as zip_ref:
                        zip_ref.extractall(tmpdirname)

                    # Conectar ao banco
                    conn = get_db_connection()
                    c = conn.cursor()

                    tabelas = ['loja', 'expedicao', 'logistica', 'comercial', 'pontuacoes']
                    for tabela in tabelas:
                        caminho_csv = os.path.join(tmpdirname, f"{tabela}.csv")
                        if os.path.exists(caminho_csv):
                            df = pd.read_csv(caminho_csv)

                            # Limpa a tabela
                            c.execute(f"DELETE FROM {tabela}")

                            # Insere os dados
                            for _, row in df.iterrows():
                                colunas = ','.join(df.columns)
                                valores = ','.join(['%s'] * len(df.columns))
                                c.execute(f"INSERT INTO {tabela} ({colunas}) VALUES ({valores})", tuple(row))

                    conn.commit()
                    conn.close()
                    flash("âœ… Backup restaurado com sucesso!", "success")
                    return redirect(url_for('home_pontuacao'))
            except Exception as e:
                flash(f"âŒ Erro ao restaurar backup: {e}", "danger")
                return redirect(url_for('restaurar_backup'))

    return render_template('restaurar_backup.html')

from io import BytesIO
from datetime import datetime
from flask import send_file
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

@app.route('/baixar_relatorio_excel')
def baixar_relatorio_excel():
    conn = get_db_connection()
    c = conn.cursor()

    tabelas = ['loja', 'expedicao', 'logistica', 'comercial', 'pontuacoes']
    nomes_formatados = {
        'loja': {
            'A': 'OrganizaÃ§Ã£o da loja (Gerente ADM)',
            'B': 'Pontualidade (RH)',
            'C': 'Fechamento do caixa (Financeiro)',
            'D': 'NÃ£o postar em rede social (RH)',
            'E': 'Validade / Avaria (Gerente ADM)'
        },
        'expedicao': {
            'A': 'OrganizaÃ§Ã£o estoque (Gerente ADM)',
            'B': 'SeparaÃ§Ã£o correta (Faturamento)',
            'C': 'Faturamento OK (Financeiro)',
            'D': 'Erros / DevoluÃ§Ãµes (Financeiro)',
            'E': 'FinalizaÃ§Ã£o apÃ³s horÃ¡rio'
        },
        'logistica': {
            'A': 'SeparaÃ§Ã£o Correta',
            'B': 'Entrega Realizada',
            'C': 'Roteiro Otimizado',
            'D': 'VeÃ­culo limpo/organizado',
            'E': 'Sem reclamaÃ§Ãµes'
        },
        'comercial': {
            'A': 'Acompanhamento pedidos',
            'B': 'ProspecÃ§Ã£o ativa',
            'C': 'Metas diÃ¡rias',
            'D': 'Ajustes manuais',
            'E': 'ParticipaÃ§Ã£o reuniÃµes'
        }
    }

    wb = Workbook()
    wb.remove(wb.active)

    try:
        for tabela in tabelas:
            # Tenta ler a tabela; se nÃ£o existir, apenas pula
            try:
                c.execute(f'SELECT * FROM {tabela}')
                dados = c.fetchall()
                colunas = [desc[0] for desc in c.description]
            except Exception:
                # tabela ausente (ou outro erro de SELECT) â†’ ignora esta aba
                conn.rollback()
                continue

            if not dados:
                continue

            ws = wb.create_sheet(title=tabela.capitalize())

            # CabeÃ§alhos amigÃ¡veis
            headers = []
            for col in colunas:
                if tabela in nomes_formatados and col.upper() in nomes_formatados[tabela]:
                    headers.append(nomes_formatados[tabela][col.upper()])
                else:
                    headers.append(col.capitalize())
            ws.append(headers)

            # Estilo do cabeÃ§alho
            for col in ws[1]:
                col.font = Font(bold=True, color="FFFFFF")
                col.fill = PatternFill(start_color="1f4e78", end_color="1f4e78", fill_type="solid")
                col.alignment = Alignment(horizontal="center", vertical="center")

            # Linhas + soma do total (tratando None)
            total_geral = 0
            idx_total = colunas.index('total') + 1 if 'total' in colunas else None

            for linha in dados:
                linha_formatada = []
                for cell in linha:
                    if cell is None or (isinstance(cell, str) and cell.lower() == 'nan'):
                        linha_formatada.append('')
                    elif isinstance(cell, datetime):
                        linha_formatada.append(cell.strftime("%d/%m/%Y"))
                    else:
                        linha_formatada.append(cell)
                ws.append(linha_formatada)

                if idx_total:
                    v = linha[colunas.index('total')]
                    try:
                        total_geral += (0 if v is None else float(v))
                    except Exception:
                        # se vier string ou algo nÃ£o numÃ©rico, ignora na soma
                        pass

            # Linha do total
            if idx_total:
                ws.append([""] * (idx_total - 1) + ["Total Geral:", total_geral])
                ultima_linha = ws.max_row
                # Coluna da palavra "Total Geral:" Ã© idx_total, o valor fica na prÃ³xima
                # RealÃ§a a cÃ©lula do rÃ³tulo
                col_letter = ws.cell(row=1, column=idx_total).column_letter
                ws[f"{col_letter}{ultima_linha}"].font = Font(bold=True, color="1f4e78")

            # Bordas e alinhamento
            thin = Side(border_style="thin", color="999999")
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
                for cell in row:
                    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
                    cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

            # Ajuste de largura
            for col in ws.columns:
                max_len = max(len(str(cell.value)) if cell.value else 0 for cell in col)
                ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 60)  # limite amigÃ¡vel

    finally:
        conn.close()

    # Envia em memÃ³ria
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(
        output,
        as_attachment=True,
        download_name=f'relatorio_pontuacoes_{datetime.now().strftime("%Y-%m-%d")}.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


@app.route('/deletar', methods=['GET', 'POST'])
def deletar():
    if request.method == 'POST':
        tabela = request.form.get('tabela')
        id_registro = request.form.get('id')
        senha = request.form.get('senha')

        if senha != DELETE_PASSWORD:
            flash("âŒ Senha incorreta.", "danger")
            return redirect(url_for('deletar'))

        if tabela not in ['loja', 'expedicao', 'logistica', 'comercial']:
            flash("âŒ Tabela invÃ¡lida.", "danger")
            return redirect(url_for('deletar'))
        try:
            conn = get_db_connection()
            c = conn.cursor()
            c.execute(f"DELETE FROM {tabela} WHERE id = %s", (id_registro,))
            conn.commit()
            conn.close()
            flash(f"âœ… Registro ID {id_registro} apagado da tabela {tabela}.", "success")
        except Exception as e:
            flash(f"âŒ Erro ao deletar: {str(e)}", "danger")

        return redirect(url_for('deletar'))

    return render_template('deletar.html')

@app.route('/ping')
def ping():
    return "OK", 200


@app.route('/admin/trigger-backup', methods=['GET', 'POST'])
def trigger_backup():
    url = fazer_backup_e_enviar()  # garantir que seja importÃ¡vel
    if url:
        return {"ok": True, "url": url}, 200
    return {"ok": False, "error": "Falha no upload. Veja logs."}, 500

# sÃ³ deixar app.run para testes locais
if __name__ == '__main__':
    app.run(debug=True)